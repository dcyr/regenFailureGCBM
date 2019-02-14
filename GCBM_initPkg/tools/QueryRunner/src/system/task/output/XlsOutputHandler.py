# core
import os
import shutil

# contrib
import win32com.client
from win32com.universal import com_error
from openpyxl.workbook import Workbook

# QueryRunner
from AbstractOutputHandler import AbstractOutputHandler
from dataimport.Cell import Cell
from system.task.output.ExportError import ExportError

class XlsOutputHandler(AbstractOutputHandler):
    '''
    Writes query output to an Excel spreadsheet (2010/.xlsx format).
    
    :param outputPath: the file to write to; if it doesn't exist, it will be
        created
    :type outputPath: str
    :param worksheet: the name of the worksheet to write query output to
    :type worksheet: str
    :param cell: optional - the cell to begin inserting the query output at
    :type cell: str
    :param header: optional - output a header row with the column names
    :type header: boolean
    '''
    
    def __init__(self, outputPath, worksheet, templatePath=None, cell="A1",
                 header=False, rotate=False, append=False):
        self.__outputPath = outputPath
        self.__templatePath = templatePath
        self.__worksheet = worksheet
        self.__cell = cell
        self.__header = header
        self.__rotate = rotate
        self.__rotation = 1
        self.__append = append


    def __writeNew(self, outPath, output):
        wb = Workbook()
        wb.save(outPath)
        
        if self.__worksheet in wb.sheetnames:
            ws = wb.get_sheet_by_name(self.__worksheet)
        else:
            ws = wb.create_sheet(title=self.__worksheet)

        if self.__header:
            ws.append(output.getColumns())
            
        for row in output:
            ws.append((value for value in row))

        wb.save(outPath)
        
        
    def __writeExisting(self, outPath, output):
        excel = None
        wb = None
           
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(outPath)
            ws = self.__getTargetWorksheet(wb)
            if not ws:
                raise ExportError("Worksheet not found: {}".format(
                                  self.__worksheet))
            
            startCell = Cell.FromString(self.__cell)
            currentRowIndex = startCell.getRowIndex() + 1
            firstColIndex = startCell.getColIndex() + 1
            
            if self.__append:
                currentRowIndex = self.__findFirstEmptyRow(ws, startCell)
            
            if self.__header:
                colNames = output.getColumns()
                cells = ws.Range(
                    ws.Cells(currentRowIndex, firstColIndex),
                    ws.Cells(currentRowIndex, firstColIndex + len(colNames) - 1))
                cells.Value = colNames
                currentRowIndex += 1
            
            for row in output:
                cells = ws.Range(
                    ws.Cells(currentRowIndex, firstColIndex),
                    ws.Cells(currentRowIndex, firstColIndex + len(row) - 1))
                cells.Value = row
                currentRowIndex += 1

            excel.CalculateFull()
            for sheet in wb.Sheets:
                for pivot in sheet.PivotTables():
                    pivot.RefreshTable()
        
        except com_error as e:
            raise ExportError("Error exporting data: {}".format(e.excepinfo[2]))
        except Exception as e:
            raise ExportError("Error in {}: {}".format(outPath, e))
        finally:
            if wb:
                wb.Close(True)
            if excel:
                excel.Quit()
                
    
    def __getTargetWorksheet(self, wb):
        for sheet in wb.Sheets:
            if sheet.Name == self.__worksheet:
                return sheet
        
        return None
                
                
    def __findFirstEmptyRow(self, ws, startCell):
        rowIndex = startCell.getRowIndex() + 1
        colIndex = startCell.getColIndex() + 1
        while True:
            cell = ws.Cells(rowIndex, colIndex)
            if cell.Value2 is None:
                return cell.Row
            rowIndex += 1

    
    def handle(self, dbTitle, output):
        '''
        See :meth:`.AbstractOutputHandler.handle`.
        '''
        if not output.getColumns():
            return
        
        outPath = self.__outputPath
        if self.__rotate:
            fileName, ext = os.path.splitext(outPath)
            outPath = "{}_{}{}".format(fileName, self.__rotation, ext)
            self.__rotation += 1
            
        if self.__templatePath and not os.path.exists(outPath):
            shutil.copyfile(self.__templatePath, outPath)
            
        if os.path.exists(outPath):
            self.__writeExisting(outPath, output)
        else:
            self.__writeNew(outPath, output)
